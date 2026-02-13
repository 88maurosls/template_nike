import io
import os
import pandas as pd
import streamlit as st
import openpyxl
from openpyxl.utils import column_index_from_string

# =========================
# STREAMLIT
# =========================
st.set_page_config(page_title="Nike Template Builder", layout="wide")
st.title("Nike Template Builder")

# =========================
# CONFIG
# =========================
TEMPLATE_PATH = "TEMPLATE NIKE.xlsx"
SOLD_TO_VALUE = 000000
SHIP_TO_VALUE = 000000

SIZE_COL_START_LETTER = "AM"
SIZE_COL_END_LETTER = "DR"

# =========================
# UI
# =========================
data_file = st.file_uploader(
    "Carica file dati (xlsx) con colonne: index, size, qty",
    type=["xlsx"]
)

# =========================
# FUNZIONI
# =========================

def clean_key(x):
    if x is None:
        return ""

    # Excel a volte trasforma 7.5 in una data tipo 05/07/2026
    if isinstance(x, (datetime.datetime, datetime.date)):
        m = x.month
        d = x.day
        # 05 -> .5
        if d == 5:
            return f"{m}.5"
        # fallback generico (se mai capitasse)
        return f"{m}.{d}".replace(",", ".")

    # numeri (int/float)
    if isinstance(x, (int, float)):
        if isinstance(x, float) and (math.isnan(x) or math.isinf(x)):
            return ""
        s = f"{x:.2f}".rstrip("0").rstrip(".")  # 7.0 -> "7", 7.5 -> "7.5"
        return s.strip().upper().replace(",", ".")

    # stringhe e altro
    return str(x).strip().upper().replace(",", ".")

def find_header_row(ws, needle="Material Number", max_scan_rows=50):
    for r in range(1, max_scan_rows + 1):
        row_values = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
        if needle in row_values:
            return r
    return None

# =========================
# MAIN
# =========================
if not data_file:
    st.info("Carica il file dati.")
    st.stop()

df = pd.read_excel(data_file)
# normalizza nomi colonne (case insensitive)
df.columns = [c.strip().lower() for c in df.columns]

required = {"index", "size", "qty"}
if not required.issubset(df.columns):
    st.error(f"Il file deve contenere: {required}")
    st.stop()

df["size_key"] = df["size"].apply(clean_key)

pivot = (
    df.pivot_table(
        index="index",
        columns="size_key",
        values="qty",
        aggfunc="sum"
    )
)

pivot.columns = [clean_key(c) for c in pivot.columns]

if not os.path.exists(TEMPLATE_PATH):
    st.error("Template non trovato.")
    st.stop()

wb = openpyxl.load_workbook(TEMPLATE_PATH)
ws = wb.active

header_row = find_header_row(ws)
if not header_row:
    st.error("Header non trovato nel template.")
    st.stop()

headers = [ws.cell(header_row, c).value for c in range(1, ws.max_column + 1)]

material_col = headers.index("Material Number") + 1
sold_to_col = headers.index("Sold To") + 1
ship_to_col = headers.index("Ship To") + 1

size_start = column_index_from_string(SIZE_COL_START_LETTER)
size_end = column_index_from_string(SIZE_COL_END_LETTER)

start_row = header_row + 1

# mapping taglie template
key_to_col = {}
for col in range(size_start, size_end + 1):
    key = clean_key(ws.cell(header_row, col).value)
    if key:
        key_to_col[key] = col

# scrittura dati
for i, sku in enumerate(pivot.index):
    r = start_row + i

    ws.cell(r, sold_to_col).value = SOLD_TO_VALUE
    ws.cell(r, ship_to_col).value = SHIP_TO_VALUE
    ws.cell(r, material_col).value = str(sku)

    for size_key, qty in pivot.loc[sku].items():
        if size_key in key_to_col:
            if pd.notna(qty) and qty > 0:
                ws.cell(r, key_to_col[size_key]).value = int(qty)

# output
out = io.BytesIO()
wb.save(out)
out.seek(0)

st.success(f"Creato file con {len(pivot.index)} SKU.")
st.download_button(
    "Scarica Excel",
    data=out.getvalue(),
    file_name="NIKE_TEMPLATE_FILLED.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
)
